#!/usr/bin/env python3
import requests
import openpyxl
from datetime import datetime

token = "API_Token"

current_time = datetime.now()
current_month = current_time.month
current_year = current_time.year
last_month = current_month - 1 if current_month > 1 else 12
target_year = current_year - 1 if current_month == 1 else current_year

start_date = datetime.strptime(f'1.{last_month}.{target_year}', '%d.%m.%Y')
end_date = datetime.strptime(f'1.{current_month}.{current_year}', '%d.%m.%Y')

start_timestamp = int(start_date.timestamp() * 1000)
end_timestamp = int(end_date.timestamp() * 1000)

url = 'https://api.opsgenie.com/v2/alerts'
headers = {"Authorization": f"GenieKey {token}"}

def get_alerts():
    params = {
        "query": f"createdAt>={start_timestamp} AND createdAt<{end_timestamp}",
        "offset": 0,
        "limit": 100,
        "sort": "createdAt",
        "order": "asc",
        "entity": True,
        "details": True
    }

    try:
        r = requests.get(url, headers=headers, params=params)
        rdata = r.json()
    except Exception as e:
        print("Erro ao buscar dados do Opsgenie.")
        print(e)
        return []

    all_data = []
    paginate = True
    while paginate:
        try:
            all_data += rdata['data']
            paging_next_url = rdata['paging'].get('next', None)
            if paging_next_url:
                r = requests.get(paging_next_url, headers=headers)
                rdata = r.json()
            else:
                paginate = False
        except Exception as e:
            print("Erro ao paginar os resultados.")
            paginate = False

    return all_data

def save_to_xlsx(data):
    if not data:
        print("Nenhum alerta encontrado para salvar.")
        return

    all_keys = set()
    flat_data = []

    for item in data:
        flat_item = {}
        for key, value in item.items():
            if isinstance(value, dict):
                for nested_key, nested_value in value.items():
                    flat_item[f"{key}_{nested_key}"] = nested_value
                    all_keys.add(f"{key}_{nested_key}")
            else:
                flat_item[key] = value
                all_keys.add(key)
        flat_data.append(flat_item)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    headers = list(all_keys)
    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_index)
        cell.value = header

    for row_index, item in enumerate(flat_data, start=2):
        for col_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = item.get(header)

    output_filename = f"alertas_opsgenie_{current_year}_{last_month}.xlsx"
    workbook.save(output_filename)
    print(f"Alertas salvos em {output_filename}")

alerts_data = get_alerts()
save_to_xlsx(alerts_data)